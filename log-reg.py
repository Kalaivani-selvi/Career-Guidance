from flask import Flask, render_template, request, redirect, session, jsonify ,url_for 
from flask_mysqldb import MySQL
import random
import string
import pandas as pd
import os
import requests
from werkzeug.utils import secure_filename

# libraries
# import numpy as np
# import pickle
# import json
# import nltk
# from keras.models import load_model
# from nltk.stem import WordNetLemmatizer

from flask import send_from_directory
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from job_roles import job_roles


app = Flask(__name__, static_url_path='/static')
app.secret_key = ''.join(random.choices(string.ascii_letters + string.digits, k=16))

## MySQL configuration
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = 'NikiKalai@2003'
app.config['MYSQL_DB'] = 'career'

mysql = MySQL(app)


## Home route
@app.route('/')
def home():
    return render_template('index.html')

## welcome route
@app.route('/welcome')
def welcome():
    return render_template('welcome.html')

# Global variable declaration
user_email = None

## Login route
@app.route('/login', methods=['GET', 'POST'])
def login():
    global user_email  # Ensure you access the global variable

    error_msg = None
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        cur = mysql.connection.cursor()
        cur.execute("SELECT * FROM users WHERE email = %s", (email,))
        user = cur.fetchone()
        cur.close()

        if user and user[2] == password:
            # Store email as global variable
            user_email = email
            # print(user_email)

            return redirect('/profile')  # Redirect to the profile page
        else:
            error_msg = 'Invalid email or password. Please try again.'

    return render_template('login.html', error_msg=error_msg)


# Mock function to fetch r1_score (replace this with your actual implementation)
def fetch_r1_score():
    # Example implementation fetching the score from the database
    cur = mysql.connection.cursor()
    cur.execute("SELECT r1_score FROM users WHERE email = %s", (user_email,))
    result = cur.fetchone()
    cur.close()
    if result:
        return result[0]
    else:
        return None
    
# Mock function to fetch r2_score
def fetch_r2_score():
    # Example implementation fetching the score from the database
    cur = mysql.connection.cursor()
    cur.execute("SELECT r2_score FROM users WHERE email = %s", (user_email,))
    result = cur.fetchone()
    cur.close()
    if result:
        return result[0]
    else:
        return None

# Mock function to fetch r3_score
def fetch_r3_score():
    # Example implementation fetching the score from the database
    cur = mysql.connection.cursor()
    cur.execute("SELECT r3_score FROM users WHERE email = %s", (user_email,))
    result = cur.fetchone()
    cur.close()
    if result:
        return result[0]
    else:
        return None

# Define the context processor to inject r2_score and r3_score into the context of all templates
@app.context_processor
def inject_scores():
    # Fetch r2_score and r3_score from the database or some other source
    r1_score = fetch_r1_score()
    r2_score = fetch_r2_score()
    r3_score = fetch_r3_score()
    return dict(r1_score=r1_score, r2_score=r2_score, r3_score=r3_score)




## Dashboard route
@app.route('/dashboard')
def dashboard():
    global user_email
    cur = mysql.connection.cursor()
    cur.execute("SELECT r1_score, r2_score, r3_score, r4_score FROM users WHERE email = %s", (user_email,))
    scores = cur.fetchone()
    cur.close()
    
    if scores:
        r1_score, r2_score, r3_score, r4_score = scores
    else:
        r1_score, r2_score, r3_score, r4_score = None, None, None, None
    
    return render_template('dashboard.html', r1_score=r1_score, r2_score=r2_score, r3_score=r3_score, r4_score=r4_score)


##pie-chart

@app.route('/get-scores')
def get_scores():
    global user_email
    cur = mysql.connection.cursor()
    cur.execute("SELECT r1_score, r2_score, r3_score, r4_score FROM users WHERE email = %s", (user_email,))
    scores = cur.fetchone()
    cur.close()

    if scores:
        r1_score, r2_score, r3_score, r4_score = scores
    else:
        r1_score, r2_score, r3_score, r4_score = 0, 0, 0, 0

    return jsonify({
        'r1_score': r1_score,
        'r2_score': r2_score,
        'r3_score': r3_score,
        'r4_score': r4_score,
    })

## Register route
@app.route('/register', methods=['GET', 'POST'])
def register():
    error_message = None  # Initialize error_message to None
    message = None  # Initialize message to None
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        # security_question = request.form['security_question']
        security_answer = request.form['security_answer']

        cur = mysql.connection.cursor()
        cur.execute("SELECT * FROM users WHERE email = %s", (email,))
        user = cur.fetchone()

        if user:
         error_message = 'Email already exists.'

        else:
            cur.execute("INSERT INTO users (email, password, security_answer) VALUES (%s, %s, %s)", (email, password, security_answer))
            mysql.connection.commit()
            cur.close()
            # Set message if registration is successful
            message = 'Registered successfully.'

    return render_template('register.html', error_message=error_message, message=message)


## Forgot password route
@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    security_error = None
    if request.method == 'POST':
        reset_email = request.form['reset_email']
        security_answer = request.form['security_answer']

        cur = mysql.connection.cursor()
        cur.execute("SELECT email FROM users WHERE email = %s AND security_answer = %s", (reset_email, security_answer))
        user_row = cur.fetchone()
        cur.close()

        if user_row:
            session['reset_email'] = reset_email
            return redirect('/reset_password')
        else:
            security_error = 'Email or security answer incorrect. Please try again.'
            

    return render_template('forgot_p.html', security_error=security_error)


## Reset password route
@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    message = None 
    if 'reset_email' in session:
        if request.method == 'POST':
            new_password = request.form['new_password']

            cur = mysql.connection.cursor()
            cur.execute("UPDATE users SET password = %s WHERE email = %s", (new_password, session['reset_email']))
            mysql.connection.commit()
            cur.close()

            session.pop('reset_email')
            message = 'Password reset successful.'


        return render_template('reset_password.html',message=message)

    return redirect('/forgot_password')

## Logout route
@app.route('/logout')
def logout():
    return render_template('index.html')


##career Guide

# Load the dataset
data = pd.read_csv("Dataset/cp.csv")
# Define job roles according to UG courses along with descriptions



@app.route('/career')
def career():
    return render_template('career_pre.html')


@app.route('/predict', methods=['POST'])
def predict_career():
    global data  # Declare data as global

    # Extract data from the HTML form
    ug_course = request.form['ugCourse']
    ug_specialization = request.form['ugSpecialization']
    cgpa = float(request.form['cgpa'])
    history_of_backlogs = int(request.form['historyofBacklogs'])
    computer_literacy = int(request.form['computerLiteracy'])
    problem_solving = int(request.form['problemSolving'])
    technical_skill = int(request.form['TechnicalSkill'])
    research_skill = int(request.form['ResearchSkill'])
    communication_skill = int(request.form['communicationSkill'])
    interpersonal_skill = int(request.form['interpersonalSkill'])
    teamwork = int(request.form['teamwork'])
    adaptability = int(request.form['adaptability'])
    time_management = int(request.form['timeManagement'])
    emotional_intelligence = int(request.form['emotionalIntelligence'])
    continuous_learning = int(request.form['continuousLearning'])
    leadership = int(request.form['leadership'])
    
    # Check if the input data exists in the dataset
    existing_data = data[(data['ugCourse'] == ug_course) & 
                         (data['ugSpecialization'] == ug_specialization) &
                         (data['cgpa'] == cgpa) &
                         (data['historyofBacklogs'] == history_of_backlogs) &
                         (data['computerLiteracy'] == computer_literacy) &
                         (data['problemSolving'] == problem_solving) &
                         (data['TechnicalSkill'] == technical_skill) &
                         (data['ResearchSkill'] == research_skill) &
                         (data['communicationSkill'] == communication_skill) &
                         (data['interpersonalSkill'] == interpersonal_skill) &
                         (data['teamwork'] == teamwork) &
                         (data['adaptability'] == adaptability) &
                         (data['timeManagement'] == time_management) &
                         (data['emotionalIntelligence'] == emotional_intelligence) &
                         (data['continuousLearning'] == continuous_learning) &
                         (data['leadership'] == leadership)]
    
    if len(existing_data) > 0:
        predicted_job_role = existing_data.iloc[0]['jobRole']
        if ug_course in job_roles and predicted_job_role in job_roles[ug_course]:
            predicted_job_summary = job_roles[ug_course][predicted_job_role]["Summary"]
            predicted_job_responsibilities = job_roles[ug_course][predicted_job_role]["Responsibilities"]
            predicted_job_qualifications = job_roles[ug_course][predicted_job_role]["Qualifications"]
            predicted_job_industry_type = job_roles[ug_course][predicted_job_role]["Industry Type"]
        else:
            predicted_job_summary = "Summary not available"
            predicted_job_responsibilities = "Responsibilities not available"
            predicted_job_qualifications = "Qualifications not available"
            predicted_job_industry_type = "Industry type not available"
    else:
        predicted_job_role = random.choice(list(job_roles[ug_course].keys()))
        predicted_job_summary = job_roles[ug_course][predicted_job_role]["Summary"]
        predicted_job_responsibilities = job_roles[ug_course][predicted_job_role]["Responsibilities"]
        predicted_job_qualifications = job_roles[ug_course][predicted_job_role]["Qualifications"]
        predicted_job_industry_type = job_roles[ug_course][predicted_job_role]["Industry Type"]

        # Append new data to dataset
        new_data = pd.DataFrame({
            "ugCourse": [ug_course],
            "ugSpecialization": [ug_specialization],
            "cgpa": [cgpa],
            "historyofBacklogs": [history_of_backlogs],
            "computerLiteracy": [computer_literacy],
            "problemSolving": [problem_solving],
            "TechnicalSkill": [technical_skill],
            "ResearchSkill": [research_skill],
            "communicationSkill": [communication_skill],
            "interpersonalSkill": [interpersonal_skill],
            "teamwork": [teamwork],
            "adaptability": [adaptability],
            "timeManagement": [time_management],
            "emotionalIntelligence": [emotional_intelligence],
            "continuousLearning": [continuous_learning],
            "leadership": [leadership],
            "jobRole": [predicted_job_role]
        })
        data = pd.concat([data, new_data], ignore_index=True)
       
    
    return render_template('result.html', predicted_job_role=predicted_job_role,
                           predicted_job_summary=predicted_job_summary,
                           predicted_job_responsibilities=predicted_job_responsibilities,
                           predicted_job_qualifications=predicted_job_qualifications,
                           predicted_job_industry_type=predicted_job_industry_type)


#chatbot


@app.route('/chatbot')
def openai():
    return render_template('openai.html')

@app.route('/chat', methods=['POST'])
def chat():
    user_message = request.form['user_message']
    response = get_chatbot_response(user_message)
    return jsonify({"message": response["result"]})

def get_chatbot_response(user_message):
    url = "https://open-ai21.p.rapidapi.com/conversationgpt35"
    payload = {
        "messages": [
            {
                "role": "user",
                "content": user_message
            }
        ],
        "web_access": False,
        "system_prompt": "",
        "temperature": 0.9,
        "top_k": 5,
        "top_p": 0.9,
        "max_tokens": 256
    }
    headers = {
        "content-type": "application/json",
        "X-RapidAPI-Key": "5167fad87bmshd240cb454662cebp106d20jsne4de16f885fd",
        "X-RapidAPI-Host": "open-ai21.p.rapidapi.com"
    }

    response = requests.post(url, json=payload, headers=headers)
    return response.json()


#chatbot

# lemmatizer = WordNetLemmatizer()


# # chat initialization
# model = load_model("Dataset/chatbot_model.h5")
# intents = json.loads(open("Intents/intents.json").read())
# words = pickle.load(open("Dataset/words.pkl", "rb"))
# classes = pickle.load(open("Dataset/classes.pkl", "rb"))


# @app.route("/chatbot")
# def chatbot():
#     return render_template("chatbot.html")


# @app.route("/get", methods=["POST"])
# def chatbot_response():
#     msg = request.form["msg"]
#     if msg.startswith('my name is'):
#         name = msg[11:]
#         ints = predict_class(msg, model)
#         if ints:
#             res1 = getResponse(ints, intents)
#             res = res1.replace("{n}", name)
#         else:
#             res = "I'm sorry, I couldn't understand that."
#     elif msg.startswith('hi my name is'):
#         name = msg[14:]
#         ints = predict_class(msg, model)
#         if ints:
#             res1 = getResponse(ints, intents)
#             res = res1.replace("{n}", name)
#         else:
#             res = "I'm sorry, I couldn't understand that."
#     else:
#         ints = predict_class(msg, model)
#         if ints:
#             res = getResponse(ints, intents)
#         else:
#             res = "I'm sorry, I couldn't understand that."
#     return res



# # chat functionalities
# def clean_up_sentence(sentence):
#     sentence_words = nltk.word_tokenize(sentence)
#     sentence_words = [lemmatizer.lemmatize(word.lower()) for word in sentence_words]
#     return sentence_words


# # return bag of words array: 0 or 1 for each word in the bag that exists in the sentence
# def bow(sentence, words, show_details=True):
#     # tokenize the pattern
#     sentence_words = clean_up_sentence(sentence)
#     # bag of words - matrix of N words, vocabulary matrix
#     bag = [0] * len(words)
#     for s in sentence_words:
#         for i, w in enumerate(words):
#             if w == s:
#                 # assign 1 if current word is in the vocabulary position
#                 bag[i] = 1
#                 if show_details:
#                     print("found in bag: %s" % w)
#     return np.array(bag)


# def predict_class(sentence, model):
#     # filter out predictions below a threshold
#     p = bow(sentence, words, show_details=False)
#     res = model.predict(np.array([p]))[0]
#     ERROR_THRESHOLD = 0.25
#     results = [[i, r] for i, r in enumerate(res) if r > ERROR_THRESHOLD]
#     # sort by strength of probability
#     results.sort(key=lambda x: x[1], reverse=True)
#     return_list = []
#     for r in results:
#         return_list.append({"intent": classes[r[0]], "probability": str(r[1])})
#     return return_list


# def getResponse(ints, intents_json):
#     tag = ints[0]["intent"]
#     list_of_intents = intents_json["intents"]
#     for i in list_of_intents:
#         if i["tag"] == tag:
#             result = random.choice(i["responses"])
#             break
#     return result


# Profile route
@app.route('/profile', methods=['GET', 'POST'])
def profile():
    global user_email  # Access the global variable
    print(user_email)

    email = user_email
    if email:
        cur = mysql.connection.cursor()
        cur.execute("SELECT * FROM users WHERE email = %s", (email,))
        user = cur.fetchone()
        cur.close()

        if user:
            password = user[2] 
            # Assuming photo_path is in the sixth column
            photo_filename = user[5]
            if photo_filename:
                photo_path = url_for('send_image', filename=photo_filename)  # Get the full path to the image
            else:
                photo_path = None
            
            name = user[4] if user[4] else None  # Assuming name is in the fifth column
            
            # Check if name and photo_path are not null
            if name is not None and photo_path is not None:
                # Set email, password, and photo_url in session
                session['email'] = email
                session['password'] = password
                session['photo_url'] = photo_path
                session['name'] = name
                
                # User details found in the database
                return render_template('profile.html', name=name, photo_path=photo_path, password=password, email=email)
            else:
                # User details not complete, show edit profile form
                return render_template('edit_profile.html', password=password, email=email)
        else:
            # User details not found, show edit profile form
            return render_template('edit_profile.html', email=email)
    else:
        # User not logged in, redirect to login page
        return redirect('/login')

    
#profile picture route
@app.route('/Pictures/<path:filename>')
def send_image(filename):
    return send_from_directory('Pictures', filename)

    
# Get the absolute path of the directory where your script resides
basedir = os.path.abspath(os.path.dirname(__file__))

# Define the upload folder relative to the directory where your script resides
app.config['UPLOAD_FOLDER'] = os.path.join(basedir, 'Pictures')


@app.route('/update_profile', methods=['GET', 'POST'])
def update_profile():
    # Retrieve form data
    name = request.form.get('uname')
    email = request.form.get('email')
   
    # Initialize the cursor outside the try block
    cur = None

    try:
        # Handle file upload
        if 'photo' in request.files:
            photo = request.files['photo']
            if photo.filename != '':
                # Save the uploaded file to a specific folder
                filename = secure_filename(photo.filename)
                photo_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                photo.save(photo_path)
                # Store the file filename in the database
                cur = mysql.connection.cursor()
                cur.execute("UPDATE users SET name=%s, photo_path=%s WHERE email=%s", (name, filename, email))
                mysql.connection.commit()

        # Update the database with the new information (excluding photo_path if no new image was provided)
        cur = mysql.connection.cursor()
        if 'photo' in request.files and request.files['photo'].filename != '':
            cur.execute("UPDATE users SET name=%s WHERE email=%s", (name, email))
        else:
            cur.execute("UPDATE users SET name=%s WHERE email=%s", (name, email))
        mysql.connection.commit()
        print("Database updated successfully")
    except Exception as e:
        # Handle database errors
        print(f"Database error: {e}")
        mysql.connection.rollback()
    finally:
        if cur:
            cur.close()

    # Redirect the user to the profile page or any other page
    return redirect('/profile')

@app.route("/edit_profile")
def edit_profile():
    # Retrieve data from the session
    email = session.get('email')
    password = session.get('password')
    photo_url = session.get('photo_url')
    name = session.get('name')
    print(photo_url)

    # Pass data to the template
    return render_template("edit_profile.html", email=email, password=password, photo_url=photo_url, name=name)


## Resume_template
@app.route('/resume')
def resume():
    return render_template('resume_selector.html')

## Rounds Start
## Round1

# Load the dataset
df = pd.read_csv("Dataset/round1.csv")

# Function to randomly select 10 questions for the quiz
def c_select_questions():
    questions_df = df.sample(n=10)
    questions_dict = questions_df.to_dict(orient='records')
    for question in questions_dict:
        options = [question['Option A'], question['Option B'], question['Option C'], question['Option D']]
        random.shuffle(options)
        question['Option A'], question['Option B'], question['Option C'], question['Option D'] = options
    return questions_dict

# Route for starting the quiz
@app.route('/c-quiz', methods=['GET', 'POST'])
def c_quiz():
    global user_email  # Ensure the global variable is accessible
    print(user_email)


    if request.method == 'GET':
        questions = c_select_questions()
        return render_template('c-quiz.html', questions=questions)
    elif request.method == 'POST':
        # Calculate score
        score = 0
        for i in range(1, 11):  # Assuming question ids are 1 to 10
            selected_option = request.form.get(f'question_{i}')
            correct_answer = df.iloc[i - 1]['Correct Answer']
            if selected_option == correct_answer:
                score += 2
        
        r1_score(user_email, score)

        
         # Check if the score is above 75% for Round 1
        if score >= 15:
            return redirect(url_for('c_result', score=score))
        else:
            message = "You need to score above 75% in Round 1 to proceed to the next round."
            return render_template('c-result.html', message=message ,score=score)


def r1_score(email, score):
    cur = mysql.connection.cursor()
    cur.execute("UPDATE users SET r1_score = %s WHERE email = %s", (score, email))
    mysql.connection.commit()
    cur.close()

# Route for displaying the result
@app.route('/c-result/<int:score>')
def c_result(score):
    return render_template('c-result.html', score=score)

## Round 2

# Load the dataset
df = pd.read_csv("Dataset/round2.csv")

# Function to randomly select 10 questions for the quiz
def a_select_questions():
    questions_df = df.sample(n=10)
    questions_dict = questions_df.to_dict(orient='records')
    for question in questions_dict:
        options = [question['Option A'], question['Option B'], question['Option C'], question['Option D']]
        random.shuffle(options)
        question['Option A'], question['Option B'], question['Option C'], question['Option D'] = options
    return questions_dict

# Route for starting the quiz
@app.route('/a-quiz', methods=['GET', 'POST'])
def a_quiz():
    global user_email  # Ensure the global variable is accessible

    # Check if the user's score in Round 1 is above 75%
    r1_score = get_r1_score(user_email)
    if r1_score is not None:
        try:
            r1_score = int(r1_score)  # Convert to integer if score is available
            if r1_score >= 15:  # Assuming maximum score is 20 (10 questions * 2 points each)
                if request.method == 'GET':
                    questions = a_select_questions()
                    return render_template('a-quiz.html', questions=questions)  # Pass r1_score to the template
                elif request.method == 'POST':
                    # Calculate score
                    score = 0
                    for i in range(1, 11):  # Assuming question ids are 1 to 10
                        selected_option = request.form.get(f'question_{i}')
                        correct_answer = df.iloc[i - 1]['Correct Answer']
                        if selected_option == correct_answer:
                            score += 2

                    r2_score(user_email, score)

                    if score >= 15:
                        return redirect(url_for('a_result', score=score))
                    else:
                        message = "You need to score above 75% in Round 2 to proceed to the next round."
                        return render_template('a-result.html', message=message ,score=score)
                    # return redirect(url_for('a_result', score=score))
            else:
                return redirect(url_for('c_quiz'))  # Redirect to Round 1 quiz if user's score in Round 1 is not above 75%
        except ValueError:
            # Handle cases where the score is not an integer
            return redirect(url_for('c_quiz'))  # Redirect to Round 1 quiz if score cannot be converted to an integer
    else:
        return redirect(url_for('c_quiz'))  # Redirect to Round 1 quiz if score is not available

# Define the function to get Round 1 score
def get_r1_score(email):
    cur = mysql.connection.cursor()
    cur.execute("SELECT r1_score FROM users WHERE email = %s", (email,))
    result = cur.fetchone()
    cur.close()
    if result:
        return result[0]
    else:
        return None

# Define the function to update Round 2 score
def r2_score(email, score):
    cur = mysql.connection.cursor()
    cur.execute("UPDATE users SET r2_score = %s WHERE email = %s", (score, email))
    mysql.connection.commit()
    cur.close()

# Route for displaying the result
@app.route('/a-result/<int:score>')
def a_result(score):
    return render_template('a-result.html', score=score)


## Round 3

df = pd.read_csv("Dataset/round3.csv")

# Function to randomly select 10 questions for the quiz
def b_select_questions():
    questions_df = df.sample(n=10)
    questions_dict = questions_df.to_dict(orient='records')
    for question in questions_dict:
        options = [question['Option A'], question['Option B'], question['Option C'], question['Option D']]
        random.shuffle(options)
        question['Option A'], question['Option B'], question['Option C'], question['Option D'] = options
    return questions_dict

# Route for starting the quiz
@app.route('/b-quiz', methods=['GET', 'POST'])
def b_quiz():
    global user_email  # Ensure the global variable is accessible

    # Check if the user's score in Round 2 is above 75%
    r2_score = get_r2_score(user_email)
    if r2_score is not None:
        try:
            r2_score = int(r2_score)  # Convert to integer if score is available
            if r2_score >= 15:  # Assuming maximum score is 20 (10 questions * 2 points each)
                if request.method == 'GET':
                    questions = b_select_questions()
                    return render_template('b-quiz.html', questions=questions)
                elif request.method == 'POST':
                    # Calculate score
                    score = 0
                    for i in range(1, 11):  # Assuming question ids are 1 to 10
                        selected_option = request.form.get(f'question_{i}')
                        correct_answer = df.iloc[i - 1]['Correct Answer']
                        if selected_option == correct_answer:
                            score += 2

                    r3_score(user_email, score)
                    if score >= 15:
                        return redirect(url_for('b_result', score=score))
                    else:
                        message = "You need to score above 75% in Round 3 to proceed to the next round."
                        return render_template('b-result.html', message=message ,score=score)
                    # return redirect(url_for('b_result', score=score))
            else:
                return redirect(url_for('a_quiz'))  # Redirect to Round 2 quiz if user's score in Round 2 is not above 75%
        except ValueError:
            # Handle cases where the score is not an integer
            return redirect(url_for('a_quiz'))  # Redirect to Round 2 quiz if score cannot be converted to an integer
    else:
        return redirect(url_for('a_quiz'))  # Redirect to Round 2 quiz if score is not available

    
# Define the function to get Round 1 score
def get_r2_score(email):
    cur = mysql.connection.cursor()
    cur.execute("SELECT r2_score FROM users WHERE email = %s", (email,))
    result = cur.fetchone()
    cur.close()
    if result:
        return result[0]
    else:
        return None


def r3_score(email, score):
    cur = mysql.connection.cursor()
    cur.execute("UPDATE users SET r3_score = %s WHERE email = %s", (score, email))
    mysql.connection.commit()
    cur.close()
# Route for displaying the result
@app.route('/b-result/<int:score>')
def b_result(score):
    return render_template('b-result.html', score=score)

##Round 4


# Load the dataset
df = pd.read_csv("Dataset/HR.csv")

# Define routes
@app.route('/hr')
def hr():
    global user_email  # Ensure the global variable is accessible

    # Check if the user's score in Round 2 is above 75%
    r3_score = get_r3_score(user_email)
    if r3_score is not None:
        try:
            r3_score = int(r3_score)  # Convert to integer if score is available
            if r3_score >= 15:  # Assuming maximum score is 20 (10 questions * 2 points each)
                roles = df['Role'].unique().tolist()
                positions = df.groupby('Role')['Position'].unique().apply(list).to_dict()
                return render_template('hr.html', roles=roles, positions=positions)
            else:
                return redirect(url_for('b_quiz'))  # Redirect to Round 2 quiz if user's score in Round 2 is not above 75%
        except ValueError:
            # Handle cases where the score is not an integer
            return redirect(url_for('b_quiz'))  # Redirect to Round 2 quiz if score cannot be converted to an integer
    else:
        return redirect(url_for('b_quiz'))  # Redirect to Round 2 quiz if score is not available





@app.route('/hr-quiz', methods=['POST'])
def hr_quiz():
    global user_email  # Ensure the global variable is accessible
    print(user_email)
    role = request.form['role']
    position = request.form['position']
    
    # Filter questions based on selected role and position
    questions_df = df[(df['Role'] == role) & (df['Position'] == position)]
    
    # Shuffle questions to randomize
    questions_df = questions_df.sample(frac=1).reset_index(drop=True)
    
    # Select the first 10 questions (or less if there are fewer than 10)
    selected_questions = questions_df.head(10)
    
    # Convert DataFrame to dictionary for easy access in the template
    questions_data = selected_questions.to_dict(orient='records')
    
    return render_template('hr-quiz.html', questions=questions_data)

# Define the function to get Round 3 score
def get_r3_score(email):
    cur = mysql.connection.cursor()
    cur.execute("SELECT r3_score FROM users WHERE email = %s", (email,))
    result = cur.fetchone()
    cur.close()
    if result:
        return result[0]
    else:
        return None

@app.route('/hr-submit', methods=['POST'])
def hr_submit():
    answers = request.form
    score = 0
    
    # Iterate over each submitted answer
    for key, submitted_answer in answers.items():
        if key.startswith('answer'):  # Check if the key is an answer
            question_index = int(key.replace('answer', ''))  # Extract the question index
            correct_answer_format = answers.get(f'correct_answer{question_index}')  # Get the correct answer format
            
            # Construct the correct answer based on the format (e.g., "a)")
            correct_answer = f"{correct_answer_format[0].lower()})"
            
            # Check if the submitted answer matches the correct answer
            if submitted_answer == correct_answer:
                score += 1  # Increment score by 1 for each correct answer
    
    # Multiply the score by 5 to get the final score out of 50
    score *= 5
    
    # Update HR round score in the database
    update_hr_score(user_email, score)
    
    return render_template('hr-result.html', score=score)

def update_hr_score(email, score):
    cur = mysql.connection.cursor()
    cur.execute("UPDATE users SET r4_score = %s WHERE email = %s", (score, email))
    mysql.connection.commit()
    cur.close()



# Function to store feedback data in an Excel file
def store_feedback(filename, name, age, q1, q2, q3, q4, q5, comments):
    timestamp = datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
    folder_path = 'Excels'
    file_path = os.path.join(folder_path, filename)
    
    try:
        # Try to load existing workbook
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        # If file doesn't exist, create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['Timestamp', 'Name', 'Age', 'Q1', 'Q2', 'Q3', 'Q4', 'Q5', 'Comments'])

    ws.append([timestamp, name, age, q1, q2, q3, q4, q5, comments])
    wb.save(file_path)

## Final feedback page route
@app.route('/f-fb')
def f_fb():
    return render_template('f-fb.html')

# Final feedback submission route
@app.route('/f-fb-submit', methods=['POST'])
def f_fb_submit():
    name = request.form['name']
    age = request.form['age']
    q1 = request.form['q1']
    q2 = request.form['q2']
    q3 = request.form['q3']
    q4 = request.form['q4']
    q5 = request.form['q5']
    comments = request.form['comments']
   
    store_feedback('final-feedback.xlsx', name, age, q1, q2, q3, q4, q5, comments)
    success_message = 'Final feedback submitted successfully!'
    return render_template('f-fb.html', success_message=success_message)

## Career Predictor Feedback page route
@app.route('/cp-fb')
def cp_fb():
    return render_template('cp-fb.html')

# Career Predictor Feedback submission route
@app.route('/cp-fb-submit', methods=['POST'])
def cp_fb_submit():
    name = request.form['cpname']
    age = request.form['cpage']
    q1 = request.form['cpq1']
    q2 = request.form['cpq2']
    q3 = request.form['cpq3']
    q4 = request.form['cpq4']
    q5 = request.form['cpq5']
    comments = request.form['comments1']

    store_feedback('cp-feedback.xlsx', name, age, q1, q2, q3, q4, q5, comments)
    success_message1 = 'Career Predictor feedback submitted successfully!'
    return render_template('cp-fb.html', success_message1=success_message1)


## ChatBot Feedback page route
@app.route('/chat-fb')
def chat_fb():
    return render_template('chat-fb.html')

# ChatBot Feedback submission route
@app.route('/chat-fb-submit', methods=['POST'])
def chat_fb_submit():
    name = request.form['cbname']
    age = request.form['cbage']
    q1 = request.form['cbq1']
    q2 = request.form['cbq2']
    q3 = request.form['cbq3']
    q4 = request.form['cbq4']
    q5 = request.form['cbq5']
    comments = request.form['comments2']

    store_feedback('chatBot-feedback.xlsx', name, age, q1, q2, q3, q4, q5, comments)
    success_message2 = 'ChatBot feedback submitted successfully!'
    return render_template('chat-fb.html', success_message2=success_message2)


## Resume Template Feedback page route
@app.route('/res-fb')
def res_fb():
    return render_template('res-fb.html')

# Resume Template submission route
@app.route('/res-fb-submit', methods=['POST'])
def res_fb_submit():
    name = request.form['resname']
    age = request.form['resage']
    q1 = request.form['resq1']
    q2 = request.form['resq2']
    q3 = request.form['resq3']
    q4 = request.form['resq4']
    q5 = request.form['resq5']
    comments = request.form['comments3']

    store_feedback('ResumeTemplate-feedback.xlsx', name, age, q1, q2, q3, q4, q5, comments)
    success_message3 = 'Resume Template feedback submitted successfully!'
    return render_template('res-fb.html', success_message3=success_message3)


## Round Feedback page route
@app.route('/r-fb')
def r_fb():
    return render_template('r-fb.html')

# Round Feedback submission route
@app.route('/r-fb-submit', methods=['POST'])
def r_fb_submit():
    name = request.form['rname']
    age = request.form['rage']
    q1 = request.form['rq1']
    q2 = request.form['rq2']
    q3 = request.form['rq3']
    q4 = request.form['rq4']
    q5 = request.form['rq5']
    comments = request.form['comments4']

    store_feedback('Rounds-feedback.xlsx', name, age, q1, q2, q3, q4, q5, comments)
    success_message4 = 'Question and Answer feedback submitted successfully!'
    return render_template('r-fb.html', success_message4=success_message4)

## Job Search page route
@app.route('/job-fb')
def job_fb():
    return render_template('job-fb.html')

# Job Search Feedback submission route
@app.route('/job-search-feedback-submit', methods=['POST'])
def job_fb_submit():
    name = request.form['jname']
    age = request.form['jage']
    q1 = request.form['jq1']
    q2 = request.form['jq2']
    q3 = request.form['jq3']
    q4 = request.form['jq4']
    q5 = request.form['jq5']
    comments = request.form['comments5']

    store_feedback('Job-Search-feedback.xlsx', name, age, q1, q2, q3, q4, q5, comments)
    success_message5 = 'Job Search feedback submitted successfully!'
    return render_template('job-fb.html', success_message5=success_message5)

##job search

# Define the API endpoint and headers
API_URL = "https://jsearch.p.rapidapi.com/search"
HEADERS = {
    "X-RapidAPI-Key": "5167fad87bmshd240cb454662cebp106d20jsne4de16f885fd",
    "X-RapidAPI-Host": "jsearch.p.rapidapi.com"
}

@app.route('/job-search', methods=['GET', 'POST'])
def job_search():
    # Default query to display peaking jobs list
    default_query = "python developer in india"
    if request.method == 'POST':
        query = request.form['query']
        # Make a request to the API with the user's query
        response = requests.get(API_URL, headers=HEADERS, params={"query": query, "page": "1", "num_pages": "1"})
        if response.status_code == 200:
            data = response.json()['data']
            return render_template('job1.html', data=data, query=query)
        else:
            return "Error fetching data from API"
    else:
        # Make a request to the API with the default query
        response = requests.get(API_URL, headers=HEADERS, params={"query": default_query, "page": "1", "num_pages": "1"})
        if response.status_code == 200:
            data = response.json()['data']
            return render_template('job1.html', data=data, query=default_query)
        else:
            return "Error fetching data from API"

if __name__ == '__main__':
    app.run(debug=True)
